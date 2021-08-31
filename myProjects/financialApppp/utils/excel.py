import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Side, Protection
from openpyxl.drawing.image import Image

from utils import database


class Excel:
    def __init__(self) -> None:
        self.dir_path = os.path.dirname(os.path.realpath(__file__))
        self.THIN = Side(border_style="thin", color="000000")
        self.DOUBLE = Side(border_style="double", color="000000")
        self.ALIGNMENT = Alignment(horizontal="center", vertical="center")
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Investimentos_Realizados"
        # Bloqueando a ABA e estabelecendo uma senha
        self.ws.protection.sheet = True
        self.ws.protection.password = '123'

        # Inserindo Imagem
        img = Image('image.jpg')
        img.anchor = 'J1'
        self.ws.add_image(img)

        # row_number = 1
        # col_idx = 10
        # cell = self.ws.cell('J1')
        # self.ws.add_image(img)

    def style_variables(self, row=0):
        # Cabeçalho Principal
        if row == 1:
            FONT = Font(bold=True, color="FFFFFF")
            PATTERN = PatternFill("solid", fgColor="00558E")
            BORDER = Border(left=self.THIN, right=self.THIN)
        # Cabeçalho Secundário
        elif row == 2:
            FONT = Font(bold=True, color="000000")
            PATTERN = PatternFill("solid", fgColor="95E9FF")
            BORDER = Border(left=self.THIN, right=self.THIN)
        # Dados dos investimentos realizados
        else:
            FONT = Font(bold=False)
            PATTERN = PatternFill("solid", fgColor="DAFFFF")
            BORDER = Border(top=self.DOUBLE, left=self.THIN,
                            right=self.THIN, bottom=self.DOUBLE)
        return FONT, PATTERN, BORDER

    def creating_excel_file(self):
        # Cabeçalho Geral
        self.ws.merge_cells('A1:F1')
        self.ws['A1'].value = 'Compras de Ativos'
        # Desbloqueia uma célula específica
        self.ws['A1'].protection = Protection(locked=False)

        # Cabeçalho secundário
        header = ['ATIVO', 'Nº COTAS', 'VALOR/COTA', 'TOTAL', 'DATA']
        self.ws.append(header)
        self.ws.merge_cells('C2:D2')
        self.ws.merge_cells('E2:F2')
        # Adquirindo dados do banco de dados
        investments = database.get_all_investments()
        i = 3  # Índice para controle das linhas
        # Variáveis para estilização das células
        (FONT, PATTERN, BORDER) = self.style_variables()
        # Ativo, Numero de cotas, Valor/cota, Total, data
        for invest in investments:
            # Armazenando dados do banco de dados em uma lista
            investment_data = [invest['ativo'], invest['numero_cotas'],
                               invest['valor_cota'], invest['numero_cotas'] *
                               invest['valor_cota'],
                               invest['data']]
            # Inserindo a lista acima no excel
            self.ws.append(investment_data)
            # Juntando as células necessárias (apenas estilização)
            self.ws.merge_cells(f"{'C'+str(i)}:{'D'+str(i)}")
            self.ws.merge_cells(f"{'E'+str(i)}:{'F'+str(i)}")
            # Estilizando a linha
            for col in range(1, 7):
                char = get_column_letter(col)
                self.ws[char+str(i)].font = FONT
                self.ws[char+str(i)].fill = PATTERN
                self.ws[char+str(i)].border = BORDER
                self.ws[char+str(i)].alignment = self.ALIGNMENT
            i += 1  # Pula uma linha

    def headers_style(self):
        (FONT, PATTERN, BORDER) = self.style_variables(1)
        for col in range(1, 7):
            char = get_column_letter(col)
            self.ws[char+'1'].font = FONT
            self.ws[char+'1'].fill = PATTERN
            self.ws[char+'1'].border = BORDER
            self.ws[char+'1'].alignment = self.ALIGNMENT

        # Cabeçalho secundário
        # Variáveis para estilização das células
        (FONT, PATTERN, BORDER) = self.style_variables(2)
        for col in range(1, 7):
            char = get_column_letter(col)
            self.ws[char+'2'].font = FONT
            self.ws[char+'2'].fill = PATTERN
            self.ws[char+'2'].border = BORDER
            self.ws[char+'2'].alignment = self.ALIGNMENT

    def save_file(self):
        excel_path = self.dir_path + 'Investments.xlsx'
        print(excel_path)
        # self.wb.save(
        #     "/mnt/c/wsl/pythonRoadMap/myProjects/financialApp/Investimentos_Teste.xlsx")
