from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from utils import database


class ExcelCreator():
    def __init__(self):
        self.work_book = load_workbook(
            '/mnt/c/wsl/pythonRoadMap/myProjects/igorProject/usuarios.xlsx')
        self.work_sheet = self.work_book.active
        # Travando as células
        self.work_sheet.protection.sheet = True
        self.work_sheet.protection.password = '123'  # Senha

    def update_excel_file(self):
        # Adicionando um cabeçalho
        self.work_sheet['A1'].value = 'Nome'
        self.work_sheet['B1'].value = 'Senha'
        self.work_sheet['C1'].value = 'e-mail'
        self.work_sheet['D1'].value = 'URL Logo'
        # Desbloqueando a célula A1
        self.work_sheet['A1'].protection = Protection(locked=False)

        # Adicionando dados do banco de dados no Excel
        users = database.get_all_users()
        # Adicionando uma lista de dados ao Excel
        for user in users:
            users_list = [user['name'], user['password'],
                          user['email'], user['logo']
                          ]
            self.work_sheet.append(users_list)

        # Setando tamanho das colunas
        for col in range(1, 5):
            self.work_sheet.column_dimensions[get_column_letter(
                col)].width = 30

    def inserting_image(self):
        img = Image('image.jpg')
        img.anchor = 'E1'
        self.work_sheet.add_image(img)

    def save_excel_file(self):  # Salva o arquivo
        self.work_book.save(
            '/mnt/c/wsl/pythonRoadMap/myProjects/igorProject/usuarios.xlsx')
